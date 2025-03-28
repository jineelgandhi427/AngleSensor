
from csv import reader

from datetime import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, PatternFill,Border,Side
from openpyxl.chart import ScatterChart, Reference,Series
import serial
import os.path

# to create workbook and take macro from path

wb = Workbook() 
wb = load_workbook(filename='F:/Thesis/Testing/Main Program cal and func test/Book1.xlsm', data_only = 'True', read_only=False, keep_vba=True)
ws = wb.worksheets[0]

# name and date in excel file name for arduino readings
    
file_path = "F:/Thesis/Testing/Reading/"
d1 = datetime.now()
#print(d1)
file_name1 = 'Calibration_run'+d1.strftime("(%Y%m%d%H%M%S)")
#print(file_name1)
fileName1=file_path + file_name1+".csv"

# name and date in excel file name for final calc

d2 = datetime.now()
#print(d2)
file_name2 = 'Calibration_run_Calc'+d2.strftime("(%Y%m%d%H%M%S)")
#print(file_name2)
fileName2=file_path + file_name2+".xlsm"

# activate arduino
ser = serial.Serial('COM14', 115200, timeout=1)


file = open(fileName1, "a")
print("Created file")
os.path.exists(fileName1)

def get_reading(ser):
  if ser.isOpen():
    
    samples = 99
    line = 0
    while line <=samples: 
      
      # selection of option 1
      result1 = ser.readline().decode('utf-8').replace("\t",",")
      ser.write(b'1')
      print(result1)
    
      # add the data to the file
      file = open(fileName1, "a", newline='')# append the data to the file
      file.write(result1) # write data with a newline

      result2 = ser.readline().decode('utf-8').replace("\t",",")
      ser.write(b'1')
      print(result2)
      
      # add the data to the file
      file = open(fileName1, "a", newline='') #append the data to the file
      file.write(result2) #write data with a newline
    

      line = line + 1
get_reading(ser)
#close out the file
file.close()

with open(fileName1, 'r') as read_obj:
    # pass the file object to reader() to get the reader object
    csv_reader = reader(read_obj)
    # Iterate over each row in the csv using reader object
    i = 1
    for row in csv_reader:
        # row variable is a list that represents a row in csv
        row_str = str(row)
        row_str = row_str.replace("[","")
        row_str =row_str.replace("]","")
        row_str = row_str.replace("'","")
        
        if i <= 17 or (i>= 149 and i<=152):
            a_cell = 'A' + str(i)
            ws[a_cell]= row_str
        if (i >17 and i<=148) or (i>=153 and i<=164) :
            arr = row_str.split(',')
            for index, column in enumerate(arr) :
                if index == 0:
                    a_cell = 'A' + str(i)
                    ws[a_cell] = column
                if index == 1:
                    b_cell = 'B' + str(i)
                    ws[b_cell] = column
                if index == 2:
                    c_cell = 'C' + str(i)
                    ws[c_cell] = column
                if index == 3:
                    d_cell = 'D' + str(i)
                    ws[d_cell] = column
                if index == 4:
                    e_cell = 'E' + str(i)
                    ws[e_cell] = column
                if index == 5:
                    f_cell = 'F' + str(i)
                    ws[f_cell] = column
        i = i+1
        
""" The styling of worksheet is considered in the next code like green color for all cell sheet, border highlighter."""
green_fill = PatternFill(fgColor='C6E0B4',patternType='solid') # same code as line 16 can be done for color
for cell in range(17,149):
    a_cell = 'A' + str(cell)
    b_cell = 'B' + str(cell)
    c_cell = 'C' + str(cell)
    d_cell = 'D' + str(cell)
    e_cell = 'E' + str(cell)
    f_cell = 'F' + str(cell)
    i_cell = 'I' + str(cell)
    j_cell = 'J' + str(cell)
    l_cell = 'L' + str(cell)
    m_cell = 'M' + str(cell)
    n_cell = 'N' + str(cell)
    o_cell = 'O' + str(cell)
    q_cell = 'Q' + str(cell)
    r_cell = 'R' + str(cell)
    t_cell = 'T' + str(cell)
    u_cell = 'U' + str(cell)
    w_cell = 'W' + str(cell)
    x_cell = 'X' + str(cell)
    z_cell = 'Z' + str(cell)
    aa_cell = 'AA' + str(cell)
    ab_cell = 'AB' + str(cell)
    ac_cell = 'AC' + str(cell)
    ad_cell = 'AD' + str(cell)
    ae_cell = 'AE' + str(cell)
    ws[a_cell].fill = green_fill
    ws[b_cell].fill = green_fill
    ws[c_cell].fill = green_fill
    ws[d_cell].fill = green_fill
    ws[e_cell].fill = green_fill
    ws[f_cell].fill = green_fill
    ws[i_cell].fill = green_fill
    ws[j_cell].fill = green_fill
    ws[l_cell].fill = green_fill
    ws[m_cell].fill = green_fill
    ws[n_cell].fill = green_fill
    ws[o_cell].fill = green_fill
    ws[q_cell].fill = green_fill
    ws[r_cell].fill = green_fill
    ws[t_cell].fill = green_fill
    ws[u_cell].fill = green_fill
    ws[w_cell].fill = green_fill
    ws[x_cell].fill = green_fill
    ws[z_cell].fill = green_fill
    ws[aa_cell].fill = green_fill
    ws[ab_cell].fill = green_fill
    ws[ac_cell].fill = green_fill
    ws[ad_cell].fill = green_fill
    ws[ae_cell].fill = green_fill

def set_colour_Mean(ws, cell_range):
    green_fill = PatternFill(fgColor='C6E0B4',patternType='solid')
    for row in ws[cell_range]:
        for cell in row:
            cell.fill = green_fill
set_colour_Mean(ws, 'I3:J4')
set_colour_Mean(ws, 'I6:J7')
set_colour_Mean(ws, 'I9:J11')
#set_colour_Mean(ws, 'A17:H148')

# The border for every cell 
def set_border(ws, cell_range):
    thin = Side(border_style="thin", color="A6A6A6")
    for row in ws[cell_range]:
        for cell in row:
            cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)
set_border(ws, 'A1:AD148') 

# bold border lines for particulare cells
def set_upper_border(ws, cell_range):
    medium = Side(border_style="medium")
    for row in ws[cell_range]:
        for cell in row:
            cell.border = Border(bottom=medium)
set_upper_border(ws, 'I2:J2') 
set_upper_border(ws, 'I4:J4') 
set_upper_border(ws, 'I5:J5') 
set_upper_border(ws, 'I7:J7') 
set_upper_border(ws, 'I8:J8') 
set_upper_border(ws, 'I11:J11') 
set_upper_border(ws, 'L2:O2') 
set_upper_border(ws, 'L7:O7') 
set_upper_border(ws, 'L8:O8') 
set_upper_border(ws, 'L13:O13') 
set_upper_border(ws, 'Q2:R2') 
set_upper_border(ws, 'Q5:R5') 
set_upper_border(ws, 'T2:U2') 
set_upper_border(ws, 'T9:U9') 

def set_right_border(ws, cell_range):
    medium = Side(border_style="medium")
    for row in ws[cell_range]:
        for cell in row:
            cell.border = Border(right=medium)
set_right_border(ws,'M3:M7')
set_right_border(ws,'M9:M13')

def set_all_border(ws, cell_range):
    medium = Side(border_style="medium")
    for row in ws[cell_range]:
        for cell in row:
            cell.border = Border(top=medium,bottom=medium,left=medium,right=medium)
set_all_border(ws, 'I17:J17')
set_all_border(ws, 'L17:O17')
set_all_border(ws, 'Q17:R17')  
set_all_border(ws, 'T17:U17')
set_all_border(ws,'W17:X17')
set_all_border(ws, 'Z17:AB17')

# Setting of Column Width for blank cell with grey color
def set_column_dimension (ws, x):
    ws.column_dimensions[x].width = 2.7
set_column_dimension(ws,'G')
set_column_dimension(ws,'H')
set_column_dimension(ws,'K')
set_column_dimension(ws,'P')
set_column_dimension(ws,'S')
set_column_dimension(ws,'V')
set_column_dimension(ws,'Y')

# Filling the cell with Grey Color
def set_color_grey (ws,cell_range):
    grey_fill = PatternFill(fgColor='DBDBDB',patternType='solid')
    for row in ws[cell_range]:
        for cell in row:
            cell.fill = grey_fill
set_color_grey(ws, 'G1:G200')
set_color_grey(ws, 'H1:H200')
set_color_grey(ws, 'K1:K200')
set_color_grey(ws, 'P1:P200')
set_color_grey(ws, 'S1:S200')
set_color_grey(ws, 'V1:V200')
set_color_grey(ws, 'Y1:Y200')

"""The naming and the unique formulas for the block I3: J11 (Mean values of offset, 
amplitude, phase and frequency)"""
ws["I3"].value = "O_X_M"
ws["I4"].value = "O_Y_M"
ws["J3"].value = "=(O6+O12)/2"
ws["J4"].value = "=(M6+M12)/2"

ws["I6"].value = "A_X_M"
ws["I7"].value = "A_Y_M"
ws["J6"].value = "=(O3+O9)/2"
ws["J7"].value = "=(M3+M9)/2"

ws["I9"].value = "ϕ_X_M"
ws["I10"].value = "ϕ_Y_M"
ws["I11"].value = "ϕ_M"
ws["J9"].value = "=(O5+O11)/2"
ws["J10"].value = "=(M5+M11)/2"
ws["J11"].value = "=J9-J10"

"""The naming and the unique formulas for the block L3: J13 (Curve fitting values offset, amplitude, phase 
and frequency)"""
ws["L3"].value = "A_Y_cw"
ws["L4"].value = "f_Y_cw"
ws["L5"].value = "ϕ_Y_cw"
ws["L6"].value = "O_Y_cw"
ws["L7"].value = "SSD"
ws["M7"].value = "=SUM(M19:M83)"

ws["N3"].value = "A_X_cw"
ws["N4"].value = "f_X_cw"
ws["N5"].value = "ϕ_X_cw"
ws["N6"].value = "O_X_cw"
ws["N7"].value = "SSD"
ws["O7"].value = "=SUM(O19:O83)"

ws["L9"].value = "A_Y_ccw"
ws["L10"].value = "f_Y_ccw"
ws["L11"].value = "ϕ_Y_ccw"
ws["L12"].value = "O_Y_ccw"
ws["L13"].value = "SSD"
ws["M13"].value = "=SUM(M84:M148)"

ws["N9"].value = "A_X_ccw"
ws["N10"].value = "f_X_ccw"
ws["N11"].value = "ϕ_X_ccw"
ws["N12"].value = "O_X_ccw"
ws["N13"].value = "SSD"
ws["O13"].value = "=SUM(O84:O148)"

"""These values are required for cure fitting calculation for the referance of GRG non-linear solver
 equation in excel file"""
ws['M3'].value = 8000
ws['M4'].value =360
ws['M5'].value = -180
ws['M6'].value =0
ws['M9'].value = 8000
ws['M10'].value =360
ws['M11'].value = -180
ws['M12'].value =0
ws['O3'].value = -8000
ws['O4'].value =360
ws['O5'].value = -180
ws['O6'].value =0
ws['O9'].value = -8000
ws['O10'].value =360
ws['O11'].value = -180
ws['O12'].value =0

"""The naming and the unique formulas for the block Q3: R5 (Positive and negative Angle Error)"""
ws. merge_cells('Q3:R3')
cell = ws.cell(row=3, column=17)
cell. value = 'Max A.E. Calibration'
cell. alignment = Alignment(horizontal='center', vertical='center')
ws["Q4"].value = "positiv AE"
ws["Q5"].value = "negativ AE"
ws["R4"].value = "=MAX(AC19:AC148)"
ws["R5"].value = "=MIN(AC19:AC148)"

"""The naming for the block T3: U8 (Note: Date,sensor, sensor_id can be typed by the user )"""
ws['T3'].value = 'Date' 
ws['T4'].value = 'Sensor'
ws['T5'].value = 'sensor_ID'
ws['T6'].value = 'Distance'
ws['U6'].value = ' mm'
ws. merge_cells('T7:T8')
cell = ws.cell(row=7, column=20)
cell. value = 'Delay zw. Microsteps'
cell. alignment = Alignment(horizontal='center', vertical='center',wrap_text=True)
ws. merge_cells('U7:U8')
cell = ws.cell(row=7, column=21)
cell. value = '250 us'
cell. alignment = Alignment(horizontal='center', vertical='center')

ws. merge_cells('I17:J17')
cell = ws.cell(row=17, column=9)
cell. value = 'Differential Output'
cell. alignment = Alignment(horizontal='center', vertical='center')

ws["I18"].value = "Y_Sin_Diff"
ws["J18"].value = "X_Cos_Diff"

ws. merge_cells('L17:O17')
cell = ws.cell(row=17, column=12)
cell. value = 'Curve Fitting'
cell. alignment = Alignment(horizontal='center', vertical='center')

ws["L18"].value = "Sin_Fit"
ws["M18"].value = "Residual^2"
ws["N18"].value = "Cos_Fit"
ws["O18"].value = "Residual^2"

ws. merge_cells('Q17:R17')
cell = ws.cell(row=17, column=17)
cell. value = 'Offset Compensation'
cell. alignment = Alignment(horizontal='center', vertical='center')

ws["Q18"].value = "Y1_Sin "
ws["R18"].value = "X1_Cos"

ws. merge_cells('T17:U17')
cell = ws.cell(row=17, column=20)
cell. value = 'Gain Compensation'
cell. alignment = Alignment(horizontal='center', vertical='center')

ws["T18"].value = "Y2_Sin "
ws["U18"].value = "X2_Cos"

ws. merge_cells('W17:X17')
cell = ws.cell(row=17, column=23)
cell. value = 'Orthogonality & Angle'
cell. alignment = Alignment(horizontal='center', vertical='center')

ws["W18"].value = "Y3_Sin"
ws["X18"].value = "alpha[deg]"

ws. merge_cells('Z17:AB17')
cell = ws.cell(row=17, column=26)
cell. value = 'Encoder'
cell. alignment = Alignment(horizontal='center', vertical='center')

ws["Z18"].value = "θ[bits]"
ws["AA18"].value = "θ[deg]"
ws["AB18"].value = "θ[rad]"
ws["AC18"].value = "Angle_Error"
ws["AD18"].value = "Hysterese"

# AE cell is used for hysterese calculation to substract the number of last row.
variable = 148
for c in range(19,84):
    ae_cell = 'AE' + str(c)
    ws[ae_cell].value = variable
    variable = variable - 1

# Column Hysterese
for i1 in range(19, 83):
    ad_cell = 'AD' + str(i1)
    ac_cell = 'AC' + str(i1)
    ae_cell = 'AE' + str(i1)
    ws[ad_cell] = '=INDIRECT("AC"&'+ae_cell+')-'+ac_cell+''
ws.formula_attributes[ad_cell]= {'t':'array', 'ref':''+ad_cell+':'+ad_cell+''}

# Every formulas which cover the row 19 to 149.
for i in  range (19,149):
    c_cell = 'C' + str(i)
    d_cell = 'D' + str(i)
    e_cell = 'E' + str(i)
    f_cell = 'F' + str(i)
    i_cell = 'I'+  str(i)
    j_cell = 'J' + str(i)
    l_cell = 'L' + str(i)
    m_cell = 'M' + str(i)
    n_cell = 'N' + str(i)
    o_cell = 'O' + str(i)
    q_cell = 'Q' + str(i)
    r_cell = 'R' + str(i)
    t_cell = 'T' + str(i)
    u_cell = 'U' + str(i)
    w_cell = 'W' + str(i)
    x_cell = 'X' + str(i)
    aa_cell = 'AA' + str(i)
    ab_cell = 'AB' + str(i)
    ac_cell = 'AC' + str(i)
    ws[i_cell] = '=' +c_cell+ '-' + e_cell +''                             # Y_Sin_Diff
    ws.formula_attributes[i_cell]= {'t':'array', 'ref':''+i_cell+':'+i_cell+''}
    ws[j_cell] = '=' +d_cell+ '-' + f_cell +''                             # X_Cos_Diff
    ws.formula_attributes[j_cell]= {'t':'array', 'ref':''+j_cell+':'+j_cell+''}
    ws[m_cell] = '=(' +i_cell+ '-' +l_cell+ ')^2'                          # Residual^2 for Sin_Fit
    ws.formula_attributes[m_cell]= {'t':'array', 'ref':''+m_cell+':'+m_cell+''}
    ws[o_cell] = '=(' +j_cell+ '-' +n_cell+ ')^2'                          # Residual_2 for Cos_Fit
    ws.formula_attributes[o_cell]= {'t':'array', 'ref':''+o_cell+':'+o_cell+''}
    ws[q_cell] = '=' +i_cell+ '-$J$4'                                      # offset compensation Y1_sin
    ws.formula_attributes[q_cell]= {'t':'array', 'ref':''+q_cell+':'+q_cell+''}
    ws[r_cell] = '=' +j_cell+ '-$J$3'                                      # offset compensation X1_cos
    ws.formula_attributes[r_cell]= {'t':'array', 'ref':''+r_cell+':'+r_cell+''}
    ws[t_cell] = '=' +q_cell+ '/$J$7'                                      # gain compensation Y2_sin
    ws.formula_attributes[t_cell]= {'t':'array', 'ref':''+t_cell+':'+t_cell+''}
    ws[u_cell] = '=' +r_cell+ '/$J$6'                                      # gain compensation X2_cos
    ws.formula_attributes[u_cell]= {'t':'array', 'ref':''+u_cell+':'+u_cell+''}
    ws[w_cell] = '=(' +t_cell+ '-' +u_cell+ '*SIN(-$J$11*PI()/180))/(COS(-$J$11*PI()/180))' # orthogonality Y3_sin
    ws.formula_attributes[w_cell]= {'t':'array', 'ref':''+w_cell+':'+w_cell+''}
    ws[ab_cell] = '='+aa_cell+'*PI()/180'                                   # thita in radian
    ws.formula_attributes[ab_cell]= {'t':'array', 'ref':''+ab_cell+':'+ab_cell+''}
    ws[ac_cell] = '=' +x_cell+'-'+aa_cell+''                                # angle error
    ws.formula_attributes[ac_cell]= {'t':'array', 'ref':''+ac_cell+':'+ac_cell+''}
    
#Sin_Fit and Cos_Fit formulas in CW direction 
for j in range(19, 84):
    aa_cell = 'AA' + str(j)
    l_cell = 'L' + str(j)
    n_cell = 'N' + str(j)
    ws[l_cell] = '=$M$3*SIN(2*PI()/$M$4*(' +aa_cell+ '+$M$5))+$M$6'         # Sin_Fit
    ws.formula_attributes[l_cell]= {'t':'array', 'ref':''+l_cell+':'+l_cell+''}
    ws[n_cell] = '=$O$3*COS(2*PI()/$O$4*(' +aa_cell+ '+$O$5))+$O$6'         #Cos_Fit
    ws.formula_attributes[n_cell]= {'t':'array', 'ref':''+n_cell+':'+n_cell+''}

#Sin_Fit and Cos_Fit formulas in CCW direction
for j in range(84, 149):
    aa_cell = 'AA' + str(j)
    l_cell = 'L' + str(j)
    n_cell = 'N' + str(j)
    ws[l_cell] = '=$M$9*SIN(2*PI()/$M$10*(' +aa_cell+ '+$M$11))+$M$12'       #Sin_Fit
    ws.formula_attributes[l_cell]= {'t':'array', 'ref':''+l_cell+':'+l_cell+''}
    ws[n_cell] = '=$O$9*COS(2*PI()/$O$10*(' +aa_cell+ '+$O$11))+$O$12'       #Cos_Fit
    ws.formula_attributes[n_cell]= {'t':'array', 'ref':''+n_cell+':'+n_cell+''}

# for alpha in degree angle
# Predefined formulas for X19 and X148
ws["X19"].value = "=IF(MOD(ATAN2(U19,W19)*180/PI()-$J$9+360,360)>5,MOD(ATAN2(U19,W19)*180/PI()-$J$9+360,360)-360,MOD(ATAN2(U19,W19)*180/PI()-$J$9+360,360))"
ws["X148"].value = '=IF(MOD(ATAN2(U148,W148)*180/PI()-$J$9+360,360)>5,MOD(ATAN2(U148,W148)*180/PI()-$J$9+360,360)-360,MOD(ATAN2(U148,W148)*180/PI()-$J$9+360,360))'

# Loop for rows 20 to 147
for k in range(20, 148):
    x_cell = 'X' + str(k)
    u_cell = 'U' + str(k)
    w_cell = 'W' + str(k)

    # Adjust the logic to match expected behavior for angles in [0, 360]
    if x_cell == 'X83':
        ws[x_cell].value = '=IF(MOD(ATAN2(U83,W83)*180/PI()-$J$9+360,360)<355,MOD(ATAN2(U83,W83)*180/PI()-$J$9+360,360)+360,MOD(ATAN2(U83,W83)*180/PI()-$J$9+360,360))'
    elif x_cell == 'X84':
        ws[x_cell].value = '=IF(MOD(ATAN2(U84,W84)*180/PI()-$J$9+360,360)<355,MOD(ATAN2(U84,W84)*180/PI()-$J$9+360,360)+360,MOD(ATAN2(U84,W84)*180/PI()-$J$9+360,360))'
    else:
        ws[x_cell].value = '=MOD(ATAN2({u},{w})*180/PI()-$J$9+360,360)'.format(u=u_cell, w=w_cell)




#Thita in bits for specific cell Z19 and Z84
ws["Z19"].value = '=I158'
ws["Z84"].value = '=I164-J158'

#Thita in bits values in CW direction
for z1 in range (20, 84):
    z_cell = 'Z' + str(z1)
    b_cell = 'B' +str(z1)

    ws[z_cell] = '='+b_cell+''
    ws.formula_attributes[z_cell]= {'t':'array', 'ref':''+z_cell+':'+z_cell+''}

#Thita in bits values in CCW direction
for z2 in range (85,149):
    z_cell = 'Z' + str(z2)
    b_cell = 'B' +str(z2)

    ws[z_cell] = '='+b_cell+''
    ws.formula_attributes[z_cell]= {'t':'array', 'ref':''+z_cell+':'+z_cell+''}

# thita in deg
ws["AA19"].value = '=360/40000*Z19'
for aa in range (20,149):
    z_cell = 'Z' + str(aa)
    aa_cell = 'AA' + str(aa)
    aa_cell1 = 'AA' +str(aa-1)

    ws[aa_cell] = '=360/40000*'+z_cell+'+'+aa_cell1+''
    ws.formula_attributes[aa_cell]= {'t':'array', 'ref':''+aa_cell+':'+aa_cell+''}

#Formulas for the encoder readings in excel sheet
ws['I152'].value = 'Start'
ws['J152'].value = 'End'
ws['I154'].value = '=0'
ws['J154'].value = '=B154'
ws['I156'].value = '=J154'
ws['J156'].value = '=I156+B156'
ws['I158'].value = '=J156'
ws['J158'].value = '=I158+B158'
ws['I160'].value = '=J158'
ws['J160'].value = '=I160+B160'
ws['I162'].value = '=J160'
ws['J162'].value = '=I162+B162'
ws['I164'].value = '=J162'
ws['J164'].value = '=I164+B164'

# The borader defination for the start and end reading of encoder in the I155: J164
def set_leftside_border(ws, cell_range):
    medium = Side(border_style="medium")
    for row in ws[cell_range]:
        for cell in row:
            cell.border = Border(left=medium)
set_leftside_border(ws, 'K152:K164')

def set_rightside_border(ws, cell_range):
    medium = Side(border_style="medium")
    for row in ws[cell_range]:
        for cell in row:
            cell.border = Border(right=medium)
set_rightside_border(ws, 'H152:H164')

medium1 = Border(top=Side(style='medium'), bottom=Side(style='medium'))
ws.cell(row=152, column=9).border = medium1
ws.cell(row=152, column=10).border = medium1
medium2 = Border( bottom=Side(style='medium'))
ws.cell(row=164, column=9).border = medium2
ws.cell(row=164, column=10).border = medium2

def set_thin_border(ws, cell_range):
    thin = Side(border_style="thin")
    for row in ws[cell_range]:
        for cell in row:
            cell.border = Border(bottom=thin)
set_thin_border(ws,'I154:J154')
set_thin_border(ws,'I156:J156')
set_thin_border(ws,'I158:J158')
set_thin_border(ws,'I160:J160')
set_thin_border(ws,'I162:J162')

""" The charts used here are the XY Scatter Chart with Smooth lines """
# Raw data (Diff. Mode) compared to the fit curve, cw
chart1 = ScatterChart()
x_value1 = Reference(ws, min_col= 27, min_row= 19, max_col=27, max_row= 83)

y_value1 = Reference(ws, min_col= 9, min_row= 19, max_col=9, max_row= 83)
series_cw1 = Series(title= 'Y_Sin_Diff', xvalues = x_value1, values= y_value1)
series_cw1.graphicalProperties.line.dashStyle = "dash"
series_cw1.graphicalProperties.line.width = 90050

y_value2 = Reference(ws, min_col= 12, min_row= 19, max_col=12, max_row= 83)
series_cw2 = Series(title= 'Sin_Fit', xvalues = x_value1, values= y_value2)
series_cw2.graphicalProperties.line.dashStyle = "solid"

y_value3 = Reference(ws, min_col= 10, min_row= 19, max_col=10, max_row= 83)
series_cw3 = Series(title= 'X_Cos_Diff', xvalues = x_value1, values= y_value3)
series_cw3.graphicalProperties.line.dashStyle = "dash"
series_cw3.graphicalProperties.line.width = 90050

y_value4 = Reference(ws, min_col= 14, min_row= 19, max_col=14, max_row= 83)
series_cw4 = Series(title= 'Cos_Fit', xvalues = x_value1, values= y_value4)
series_cw4.graphicalProperties.line.dashStyle = "solid"

chart1.series.append(series_cw1)
chart1.series.append(series_cw2)
chart1.series.append(series_cw3)
chart1.series.append(series_cw4)
chart1.legend.position= 'b'
chart1.title = " Raw data (Diff. Mode) compared to fit curve, cw "
ws.add_chart(chart1,'L150')

# Raw data (Diff. Mode) compared to the fit curve, ccw
chart2 = ScatterChart()
x_value2 = Reference(ws, min_col= 27, min_row= 84, max_col=27, max_row= 148)

y_value_ccw1 = Reference(ws, min_col= 9, min_row= 84, max_col=9, max_row= 148)
series_ccw1 = Series(title= 'Y_Sin_Diff', xvalues = x_value2, values= y_value_ccw1)
series_ccw1.graphicalProperties.line.dashStyle = "dash"
series_ccw1.graphicalProperties.line.width = 90050

y_value_ccw2 = Reference(ws, min_col= 12, min_row= 84, max_col=12, max_row= 148)
series_ccw2 = Series(title= 'Sin_Fit', xvalues = x_value2, values= y_value_ccw2)
series_ccw2.graphicalProperties.line.dashStyle = "solid"

y_value_ccw3 = Reference(ws, min_col= 10, min_row= 84, max_col=10, max_row= 148)
series_ccw3 = Series(title= 'X_Cos_Diff', xvalues = x_value2, values= y_value_ccw3)
series_ccw3.graphicalProperties.line.dashStyle = "dash"
series_ccw3.graphicalProperties.line.width = 90050

y_value_ccw4 = Reference(ws, min_col= 14, min_row= 84, max_col=14, max_row= 148)
series_ccw4 = Series(title= 'Cos_Fit', xvalues = x_value2, values= y_value_ccw4)
series_ccw4.graphicalProperties.line.dashStyle = "solid"

chart2.series.append(series_ccw1)
chart2.series.append(series_ccw2)
chart2.series.append(series_ccw3)
chart2.series.append(series_ccw4)
chart2.legend.position= 'b'
chart2.title = "Raw data (Diff. Mode) compared to fit curve, ccw"
ws.add_chart(chart2,'L165')

#Angle error & hysteresis calibration run graph
chart3 = ScatterChart()

x_value_CW = Reference(ws, min_col= 27, min_row= 19, max_col=27, max_row= 83)
x_value_CCW = Reference(ws, min_col= 27, min_row= 84, max_col=27, max_row= 148)

y_value_CW = Reference(ws, min_col= 29, min_row= 19, max_col=29, max_row= 83)
series_CW = Series(title= 'cw', xvalues = x_value_CW, values= y_value_CW)
series_CW.graphicalProperties.line.dashStyle = "solid"

y_value_CCW = Reference(ws, min_col= 29, min_row= 84, max_col=29, max_row= 148)
series_CCW = Series(title= 'ccw', xvalues = x_value_CCW, values= y_value_CCW)
series_CCW.graphicalProperties.line.dashStyle = "solid"

y_value_hys = Reference(ws, min_col= 30, min_row= 19, max_col=30, max_row= 83)
series_hys = Series(title= 'hysterese', xvalues = x_value_CW, values= y_value_hys)
series_hys.graphicalProperties.line.dashStyle = "solid"

chart3.series.append(series_CW)
chart3.series.append(series_CCW)
chart3.series.append(series_hys)

chart3.legend.position= 'b'
chart3.title = "Angle error & hysteresis calibration run"
ws.add_chart(chart3,'X150')

wb.save(fileName2)











